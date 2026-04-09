# apps/statistic/exports.py

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Avg
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from apps.patients.models import PatientCard


def get_filtered_queryset(request):
    from apps.users.decorators import department_filter

    qs = PatientCard.objects.select_related(
        'department', 'attending_doctor', 'referral_organization',
        'country', 'region', 'district', 'city',
        'discharge_conclusion'
    ).prefetch_related(
        'operations__operation_type'
    ).order_by('admission_date')

    # Rol bo'yicha cheklash
    qs = department_filter(qs, request.user)

    # Barcha filterlar
    year = request.GET.get('year')
    month = request.GET.get('month')
    dept = request.GET.get('department')
    doctor = request.GET.get('doctor')
    outcome = request.GET.get('outcome')
    status = request.GET.get('status')
    gender = request.GET.get('gender')
    patient_category = request.GET.get('patient_category')
    resident_status = request.GET.get('resident_status')
    referral_type = request.GET.get('referral_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if year:
        qs = qs.filter(admission_date__year=year)
    if month:
        qs = qs.filter(admission_date__month=month)
    if dept:
        qs = qs.filter(department_id=dept)
    if doctor:
        qs = qs.filter(attending_doctor_id=doctor)
    if outcome:
        qs = qs.filter(outcome=outcome)
    if status:
        qs = qs.filter(status=status)
    if gender:
        qs = qs.filter(gender=gender)
    if patient_category:
        qs = qs.filter(patient_category=patient_category)
    if resident_status:
        qs = qs.filter(resident_status=resident_status)
    if referral_type:
        qs = qs.filter(referral_type=referral_type)
    if date_from:
        qs = qs.filter(admission_date__date__gte=date_from)
    if date_to:
        qs = qs.filter(admission_date__date__lte=date_to)

    from datetime import date as dt_date, timedelta
    age_group = request.GET.get('age_group', '')
    if age_group == 'under16':
        from datetime import date
        from dateutil.relativedelta import relativedelta
        # Python darajasida 16 yoshgacha filtrlash
        today = date.today()
        cutoff = today - relativedelta(years=16)  # 16 yil oldingi sana
        qs = qs.filter(birth_date__isnull=False, birth_date__gt=cutoff)
    elif age_group == 'adult':
        from datetime import date
        from dateutil.relativedelta import relativedelta
        today = date.today()
        cutoff = today - relativedelta(years=16)
        qs = qs.filter(birth_date__isnull=False, birth_date__lte=cutoff)

    org_id = request.GET.get('org')
    if org_id:
        qs = qs.filter(workplace_org_id=org_id)

    return qs


def get_stats_queryset(request):
    """Statistika uchun alohida queryset — prefetch yo'q"""
    qs = PatientCard.objects.all()
    year = request.GET.get('year')
    dept = request.GET.get('department')
    if year:
        qs = qs.filter(admission_date__year=year)
    if dept:
        qs = qs.filter(department_id=dept)
    return qs


def get_full_address(patient):
    parts = filter(None, [
        str(patient.country) if patient.country else '',
        str(patient.region) if patient.region else '',
        str(patient.district) if patient.district else '',
        str(patient.city) if patient.city else '',
        patient.street_address or '',
    ])
    return ', '.join(parts) or '—'


def export_excel(request):
    qs = get_filtered_queryset(request)
    qs_stats = get_stats_queryset(request)

    wb = openpyxl.Workbook()

    # ==================== STIL O'ZGARUVCHILARI ====================
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ==================== 1-SAHIFA: BEMORLAR RO'YXATI ====================
    ws = wb.active
    ws.title = "Bemorlar ro'yxati"

    headers = [
        "№", "Bayonnoma №", "Ism-familiya", "Jinsi", "Tug'ilgan sana",
        "Rezident", "Manzil", "Ijtimoiy holat", "Passport", "Qabul turi",
        "Yo'llagan muassasa", "Qabul bo'limi tashxisi", "Necha soat keyin",
        "Shoshilinch", "Pullik", "Yotqizilgan sana", "Bo'lim",
        "Qayta/Birinchi", "Yotgan kunlar", "Yakun", "Chiqish xulosasi",
        "Chiqgan sana", "Klinik tashxis (MKB-10)", "Shifokor",
        "Jarrohlik amaliyotlari",
    ]

    ws.row_dimensions[1].height = 40
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_num, patient in enumerate(qs, 2):
        # Jarrohlik amaliyotlarini yig'ish
        operations = patient.operations.all()
        if operations:
            op_lines = []
            for i, op in enumerate(operations, 1):
                name = str(op.operation_type) if op.operation_type else op.operation_name or '—'
                date = op.operation_date.strftime('%d.%m.%Y') if op.operation_date else '—'
                anesthesia = op.get_anesthesia_display() if op.anesthesia else '—'
                complication = op.complication or '—'
                op_lines.append(
                    f"{i}. {date} | {name} | Narkoz: {anesthesia} | Asorat: {complication}"
                )
            operations_text = '\n'.join(op_lines)
        else:
            operations_text = '—'

        row_data = [
            row_num - 1,
            patient.medical_record_number,
            patient.full_name,
            patient.get_gender_display(),
            patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '',
            patient.get_resident_status_display(),
            get_full_address(patient),
            patient.get_social_status_display(),
            patient.passport_serial or '',
            patient.get_referral_type_display(),
            str(patient.referral_organization) if patient.referral_organization else '',
            patient.admission_diagnosis or '',
            patient.get_hours_after_illness_display(),
            'Ha' if patient.is_emergency else "Yo'q",
            'Ha' if patient.is_paid else "Yo'q",
            patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '',
            str(patient.department) if patient.department else '',
            patient.get_admission_count_display(),
            patient.days_in_hospital,
            patient.get_outcome_display(),
            str(patient.discharge_conclusion) if patient.discharge_conclusion else '',
            patient.discharge_date.strftime('%d.%m.%Y %H:%M') if patient.discharge_date else '',
            patient.clinical_main_diagnosis or '',
            str(patient.attending_doctor) if patient.attending_doctor else '',
            operations_text,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border

            # Yakun bo'yicha rang
            if col_num == 20:
                if value == 'Chiqarildi':
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value == 'Vafot etdi':
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                elif value == "Boshqa shifoxonaga o'tkazildi":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")

    # Ustun kengliklari
    col_widths = [
        4, 14, 25, 8, 13, 10, 30, 15, 12, 16,
        22, 25, 18, 11, 8, 18, 18, 14, 12, 14,
        18, 18, 16, 22, 40
    ]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Jarrohlik ustuni qator balandligi
    for row_num in range(2, ws.max_row + 1):
        ops_cell = ws.cell(row=row_num, column=25)
        if ops_cell.value and '\n' in str(ops_cell.value):
            line_count = str(ops_cell.value).count('\n') + 1
            ws.row_dimensions[row_num].height = max(15, line_count * 15)

    # ==================== 2-SAHIFA: STATISTIKA ====================
    ws2 = wb.create_sheet("Statistika")

    ws2['A1'] = "Umumiy statistika"
    ws2['A1'].font = Font(bold=True, size=14)

    stats = [
        ("Jami bemorlar", qs_stats.count()),
        ("Chiqarildi", qs_stats.filter(outcome='discharged').count()),
        ("Vafot etdi", qs_stats.filter(outcome='deceased').count()),
        ("O'tkazildi", qs_stats.filter(outcome='transferred').count()),
        ("Shoshilinch qabul", qs_stats.filter(is_emergency=True).count()),
        ("Pullik yotqizilgan", qs_stats.filter(is_paid=True).count()),
        ("Rezidentlar", qs_stats.filter(resident_status='resident').count()),
        ("Norezidentlar", qs_stats.filter(resident_status='non_resident').count()),
        ("O'rtacha yotish kunlari", round(
            qs_stats.aggregate(avg=Avg('days_in_hospital'))['avg'] or 0, 1
        )),
    ]

    for row_num, (label, value) in enumerate(stats, 3):
        label_cell = ws2.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True)
        ws2.cell(row=row_num, column=2, value=value)

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15

    # ==================== 3-SAHIFA: JARROHLIK AMALIYOTLARI ====================
    ws3 = wb.create_sheet("Jarrohlik amaliyotlari")

    op_headers = [
        "№", "Bemor", "Bayonnoma №", "Sana",
        "Amaliyot nomi", "Narkoz", "Asorati"
    ]
    for col_num, header in enumerate(op_headers, 1):
        cell = ws3.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    op_row = 2
    for patient in qs:
        for op in patient.operations.all():
            name = str(op.operation_type) if op.operation_type else op.operation_name or '—'
            op_data = [
                op_row - 1,
                patient.full_name,
                patient.medical_record_number,
                op.operation_date.strftime('%d.%m.%Y') if op.operation_date else '—',
                name,
                op.get_anesthesia_display() if op.anesthesia else '—',
                op.complication or '—',
            ]
            for col_num, value in enumerate(op_data, 1):
                cell = ws3.cell(row=op_row, column=col_num, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
            op_row += 1

    if op_row == 2:
        ws3.cell(row=2, column=1, value="Jarrohlik amaliyotlari yo'q")

    op_col_widths = [4, 25, 14, 12, 35, 15, 25]
    for i, width in enumerate(op_col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    # ==================== 4-SAHIFA: BEMORLAR + XIZMATLAR ====================
    from apps.services.models import PatientService
    from django.db.models import Sum, Count

    # ==================== 4-SAHIFA: BEMORLAR XIZMATLARI ====================
    from apps.services.models import PatientService, PatientMedicine
    from django.db.models import Sum, Count

    ws4 = wb.create_sheet("Bemorlar xizmatlari")

    # Ustun sarlavhalari — bemor ma'lumotlari + xizmatlar + dorilar
    BEM_HEADS = [
        '№', 'F.I.Sh', 'Qabul sanasi va vaqt', "Tug'ilgan sana",
        'Passport', 'JSHSHIR', 'Yashash manzili', 'Lavozimi', 'Ish joyi',
        'Ota-ona ismi', 'Ota-ona JSHSHIR', 'Ota-ona ish joyi',
        'Qabulxona tashxisi', "Bo'lim", 'Bemor turi',
        'Yotoq kun', 'Yil boshidan', 'Jami tashriflar',
        'Xizmat turi', "Xizmat summasi (so'm)",
        'Dori nomi', "Dori summasi (so'm)",
    ]

    # Ustun kengliklari
    col_widths_4 = [5, 28, 18, 14, 14, 16, 30, 18, 28, 24, 16, 26, 30, 20, 14, 10, 12, 14, 25, 18, 25, 18]
    for ci, w in enumerate(col_widths_4, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BEM_HEADS))
    c = ws4.cell(row=1, column=1, value="BEMORLAR XIZMATLARI VA DORI-DARMONLAR HISOBOTI")
    c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    ws4.row_dimensions[1].height = 28

    for col, h in enumerate(BEM_HEADS, 1):
        c = ws4.cell(row=2, column=col, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    ws4.row_dimensions[2].height = 30

    r4 = 3
    patient_num = 0
    grand_svc_total = 0
    grand_med_total = 0
    grand_total_4   = 0

    from django.db.models import Q
    from datetime import date

    year_start = date(date.today().year, 1, 1)

    for patient in qs:
        p_svcs  = PatientService.objects.filter(
            patient_card=patient
        ).select_related('service__category').order_by('service__category__name')
        p_meds  = PatientMedicine.objects.filter(
            patient_card=patient
        ).select_related('medicine').order_by('medicine__name')

        has_svc = p_svcs.exists()
        has_med = p_meds.exists()
        if not has_svc and not has_med:
            continue

        patient_num += 1

        # Bemor umumiy ma'lumotlari
        address_parts = [
            str(patient.country) if patient.country else '',
            str(patient.region)  if patient.region  else '',
            str(patient.district)if patient.district else '',
            str(patient.city)    if patient.city     else '',
            patient.street_address or '',
        ]
        address = ', '.join(p for p in address_parts if p) or '—'

        workplace = ''
        if patient.workplace_org:
            workplace = str(patient.workplace_org)
        elif patient.workplace:
            workplace = patient.workplace

        # Ota-ona ma'lumotlari (16 yoshgacha temir yo'lchi bolalar uchun)
        parent_name_val = patient.parent_name or '—'
        parent_jshshir_val = patient.parent_jshshir or '—'
        parent_org_val = '—'
        if hasattr(patient, 'parent_workplace_org') and patient.parent_workplace_org:
            parent_org_val = str(patient.parent_workplace_org)

        # Tashriflar soni
        from apps.patients.models import PatientCard as PC
        visits_year  = PC.objects.filter(
            full_name=patient.full_name,
            admission_date__date__gte=year_start
        ).count() if patient.full_name else 0
        visits_total = PC.objects.filter(
            full_name=patient.full_name
        ).count() if patient.full_name else 0
        if patient.JSHSHIR:
            visits_year  = PC.objects.filter(JSHSHIR=patient.JSHSHIR, admission_date__date__gte=year_start).count()
            visits_total = PC.objects.filter(JSHSHIR=patient.JSHSHIR).count()

        # Xizmatlar kategoriya bo'yicha
        cat_groups = {}
        for svc in p_svcs:
            cat = svc.service.category.name
            cat_groups.setdefault(cat, 0)
            cat_groups[cat] += float(svc.price * svc.quantity)

        # Dori-darmonlar
        med_groups = {}
        for med in p_meds:
            name = med.medicine.name
            med_groups.setdefault(name, 0)
            med_groups[name] += float(med.total_price)

        # Nechta qator kerak
        max_rows = max(len(cat_groups) if cat_groups else 1,
                       len(med_groups) if med_groups else 1)

        svc_items = list(cat_groups.items())
        med_items = list(med_groups.items())

        # Har bir qator
        for ri in range(max_rows):
            row_num_in_patient = ri

            # Bemor info — faqat birinchi qatorda
            if ri == 0:
                bem_data = [
                    patient_num,
                    patient.full_name,
                    patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '—',
                    patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '—',
                    patient.passport_serial or '—',
                    patient.JSHSHIR or '—',
                    address,
                    patient.position or '—',
                    workplace or '—',
                    parent_name_val,
                    parent_jshshir_val,
                    parent_org_val,
                    patient.admission_diagnosis or '—',
                    str(patient.department) if patient.department else '—',
                    patient.get_patient_category_display(),
                    patient.days_in_hospital or 0,
                    visits_year,
                    visits_total,
                ]
            else:
                bem_data = [''] * 15

            # Xizmat
            if ri < len(svc_items):
                cat_name, cat_sum = svc_items[ri]
                bem_data += [cat_name, cat_sum]
            else:
                bem_data += ['', '']

            # Dori
            if ri < len(med_items):
                med_name, med_sum = med_items[ri]
                bem_data += [med_name, med_sum]
            else:
                bem_data += ['', '']

            for col, val in enumerate(bem_data, 1):
                c = ws4.cell(row=r4, column=col, value=val)
                c.border = border
                c.font = Font(size=9)
                c.alignment = Alignment(
                    horizontal='center' if col in (1,3,4,13,14,15) else 'left',
                    vertical='center', wrap_text=True
                )
                if col in (17, 19): c.number_format = '#,##0'
                if ri % 2 == 0:
                    c.fill = PatternFill('solid', fgColor='F8F9FA')

            ws4.row_dimensions[r4].height = 18
            r4 += 1

        # Bemor jami qatori
        svc_total  = sum(v for _, v in svc_items)
        med_total_ = sum(v for _, v in med_items)
        bem_jami   = svc_total + med_total_

        # Jami qatorni birlashtirish (1-15 ustunlar)
        if max_rows > 1:
            ws4.merge_cells(
                start_row=r4 - max_rows, start_column=1,
                end_row=r4 - 1, end_column=1
            )

        jami_row_data = [''] * 15 + ['Xizmatlar jami:', svc_total, 'Dorilar jami:', med_total_]
        for col, val in enumerate(jami_row_data, 1):
            c = ws4.cell(row=r4, column=col, value=val)
            c.border = border
            c.font = Font(size=9, bold=True)
            c.fill = PatternFill('solid', fgColor='D6E4F0')
            c.alignment = Alignment(horizontal='right' if col in (16,18) else 'center', vertical='center')
            if col in (17, 19): c.number_format = '#,##0'
        # Bemor umumiy jami (19-ustundan keyin alohida)
        ws4.cell(row=r4, column=1).value = "Bemor jami: {:,.0f} so'm".format(bem_jami)
        ws4.cell(row=r4, column=1).font = Font(size=9, bold=True)
        ws4.cell(row=r4, column=1).fill = PatternFill('solid', fgColor='D6E4F0')
        ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=15)
        ws4.row_dimensions[r4].height = 18
        r4 += 1

        # Umumiy jami uchun yig'amiz
        grand_svc_total  += svc_total
        grand_med_total  += med_total_
        grand_total_4    += bem_jami

    # Barcha bemorlar umumiy jami
    ws4.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=15)
    c = ws4.cell(row=r4, column=1, value="BARCHA BEMORLAR UMUMIY JAMI: {:,.0f} so'm".format(grand_total_4))
    c.fill = header_fill; c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    ws4.cell(row=r4, column=16, value="Xizmatlar jami:").border = border
    c17 = ws4.cell(row=r4, column=17, value=grand_svc_total)
    c17.fill = header_fill; c17.font = header_font
    c17.number_format = '#,##0'; c17.border = border
    ws4.cell(row=r4, column=18, value="Dorilar jami:").border = border
    c19 = ws4.cell(row=r4, column=19, value=grand_med_total)
    c19.fill = header_fill; c19.font = header_font
    c19.number_format = '#,##0'; c19.border = border
    ws4.row_dimensions[r4].height = 24

    # ==================== 5-SAHIFA: BEMOR + XIZMATLAR ====================
    from apps.services.models import ServiceCategory

    ws5 = wb.create_sheet("Bemor + Xizmatlar")

    all_cats = list(ServiceCategory.objects.filter(is_active=True).order_by('name'))
    cat_names = [c.name for c in all_cats]

    # Ustunlar: bemor ma'lumotlari + kategoriyalar + dorilar jami + JAMI
    FIXED = [
        '№', 'F.I.Sh', 'Qabul sanasi', "Tug'ilgan sana",
        'Passport', 'JSHSHIR', 'Yashash manzili', 'Lavozimi', 'Ish joyi',
        'Ota-ona ismi', 'Ota-ona JSHSHIR', 'Ota-ona ish joyi',
        'Qabulxona tashxisi', "Bo'lim", 'Bemor turi',
        'Yotoq kun', 'Yil boshidan', 'Jami tashriflar',
    ]
    all_headers = FIXED + cat_names + ["Dori-darmonlar (so'm)", "JAMI (so'm)"]
    total_cols5 = len(all_headers)

    # Ustun kengliklari
    fixed_widths = [5, 28, 16, 14, 14, 16, 28, 16, 25, 22, 14, 24, 28, 18, 14, 10, 12, 14]
    ws5.column_dimensions['A'].width = 5
    for ci, w in enumerate(fixed_widths, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(len(FIXED)+1, len(FIXED)+len(cat_names)+1):
        ws5.column_dimensions[get_column_letter(ci)].width = 16
    ws5.column_dimensions[get_column_letter(len(all_headers)-1)].width = 18
    ws5.column_dimensions[get_column_letter(len(all_headers))].width = 18

    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols5)
    c = ws5.cell(row=1, column=1, value="BEMORLAR VA XIZMATLAR — KENGAYTIRILGAN JADVAL")
    c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    ws5.row_dimensions[1].height = 28

    for col, h in enumerate(all_headers, 1):
        c = ws5.cell(row=2, column=col, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    ws5.row_dimensions[2].height = 35

    r5 = 3
    col_grand_totals = {name: 0 for name in cat_names}
    overall_total5   = 0
    med_grand_total  = 0

    for num, patient in enumerate(qs, 1):
        # Xizmatlar
        svc_by_cat = {}
        for row in PatientService.objects.filter(patient_card=patient).values(
            'service__category__name'
        ).annotate(total=Sum('price')):
            svc_by_cat[row['service__category__name']] = float(row['total'] or 0)

        # Dorilar jami
        med_total = float(
            PatientMedicine.objects.filter(patient_card=patient)
            .aggregate(t=Sum('price'))[ 't'] or 0
        )

        patient_svc_total = sum(svc_by_cat.values())
        patient_total     = patient_svc_total + med_total
        overall_total5   += patient_total
        med_grand_total  += med_total

        # Bemor ma'lumotlari
        address_parts = [
            str(patient.country) if patient.country else '',
            str(patient.region)  if patient.region  else '',
            str(patient.district)if patient.district else '',
            str(patient.city)    if patient.city     else '',
            patient.street_address or '',
        ]
        address = ', '.join(p for p in address_parts if p) or '—'
        workplace = str(patient.workplace_org) if patient.workplace_org else (patient.workplace or '—')

        if patient.JSHSHIR:
            vis_year  = PatientCard.objects.filter(JSHSHIR=patient.JSHSHIR, admission_date__date__gte=year_start).count()
            vis_total = PatientCard.objects.filter(JSHSHIR=patient.JSHSHIR).count()
        else:
            vis_year = vis_total = 1

        row_data = [
            num,
            patient.full_name,
            patient.admission_date.strftime('%d.%m.%Y') if patient.admission_date else '—',
            patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '—',
            patient.passport_serial or '—',
            patient.JSHSHIR or '—',
            address,
            patient.position or '—',
            workplace,
            patient.parent_name or '—',
            patient.parent_jshshir or '—',
            str(patient.parent_workplace_org) if getattr(patient, 'parent_workplace_org', None) else '—',
            patient.admission_diagnosis or '—',
            str(patient.department) if patient.department else '—',
            patient.get_patient_category_display(),
            patient.days_in_hospital or 0,
            vis_year,
            vis_total,
        ]

        for cat_name in cat_names:
            val = svc_by_cat.get(cat_name, 0)
            row_data.append(val if val else '')
            col_grand_totals[cat_name] = col_grand_totals.get(cat_name, 0) + val

        row_data.append(med_total if med_total else '')
        row_data.append(patient_total if patient_total else '')

        for col, val in enumerate(row_data, 1):
            c = ws5.cell(row=r5, column=col, value=val)
            c.border = border
            c.font = Font(size=9)
            c.alignment = Alignment(
                horizontal='center' if col in (1,3,4,13,14,15) else 'left',
                vertical='center', wrap_text=True
            )
            if col > len(FIXED): c.number_format = '#,##0'
            if num % 2 == 0: c.fill = PatternFill('solid', fgColor='F8F9FA')
        ws5.row_dimensions[r5].height = 18
        r5 += 1

    # Jami qator
    jami_row = ['JAMI:'] + [''] * (len(FIXED)-1)
    for cat_name in cat_names:
        jami_row.append(col_grand_totals.get(cat_name, 0))
    jami_row.append(med_grand_total)
    jami_row.append(overall_total5)

    for col, val in enumerate(jami_row, 1):
        c = ws5.cell(row=r5, column=col, value=val)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal='center' if col==1 else 'right', vertical='center')
        c.border = border
        if col > 1 and isinstance(val, (int, float)): c.number_format = '#,##0'
    ws5.row_dimensions[r5].height = 22

    # ==================== 6-SAHIFA: TASHKILOT BO'YICHA ====================
    from apps.patients.models import Organization
    from apps.services.models import PatientService

    ws6 = wb.create_sheet("Tashkilotlar")
    ws6.column_dimensions['A'].width = 5
    ws6.column_dimensions['B'].width = 40
    ws6.column_dimensions['C'].width = 12
    ws6.column_dimensions['D'].width = 12
    ws6.column_dimensions['E'].width = 12
    ws6.column_dimensions['F'].width = 20
    ws6.column_dimensions['G'].width = 20
    ws6.column_dimensions['H'].width = 20

    # Sarlavha
    ws6.merge_cells('A1:H1')
    c = ws6.cell(row=1, column=1, value="TASHKILOT BO'YICHA STATISTIKA")
    c.fill = header_fill; c.font = header_font
    c.alignment = center; c.border = border
    ws6.row_dimensions[1].height = 28

    # Ustun sarlavhalar
    h6 = [
        '№', 'Tashkilot', 'Korxona kodi', 'Filial kodi',
        'Bemorlar', "Bo'lim bo'yicha", 'Tashxis', "Xizmatlar summasi (so'm)"
    ]
    for col, h in enumerate(h6, 1):
        c = ws6.cell(row=2, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = center; c.border = border
    ws6.row_dimensions[2].height = 25

    # Tashkilotlar
    org_qs = (
        qs.filter(workplace_org__isnull=False)
        .values(
            'workplace_org__id',
            'workplace_org__enterprise_name',
            'workplace_org__branch_name',
            'workplace_org__enterprise_code',
            'workplace_org__branch_code',
        )
        .annotate(patient_count=Count('id'))
        .order_by('-patient_count')
    )

    r6 = 3
    grand_patients = 0
    grand_svc_total = 0

    for num, org in enumerate(org_qs, 1):
        org_id = org['workplace_org__id']
        ent_name = org['workplace_org__enterprise_name'] or ''
        branch   = org['workplace_org__branch_name'] or ''
        full_name = f"{ent_name} — {branch}" if branch else ent_name
        p_count  = org['patient_count']
        grand_patients += p_count

        # Xizmatlar summasi
        svc_total = PatientService.objects.filter(
            patient_card__workplace_org_id=org_id
        ).filter(
            patient_card__in=qs
        ).aggregate(t=Sum('price'))['t'] or 0
        svc_total = float(svc_total)
        grand_svc_total += svc_total

        # Bo'lim taqsimoti
        dept_dist = list(
            qs.filter(workplace_org_id=org_id)
            .values('department__name')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')[:3]
        )
        dept_str = ', '.join([
            f"{d['department__name']}({d['cnt']})"
            for d in dept_dist if d['department__name']
        ]) or '—'

        # Tashxis (eng ko'p uchraydigan)
        diag_dist = list(
            qs.filter(workplace_org_id=org_id)
            .exclude(admission_diagnosis='')
            .values('admission_diagnosis')
            .annotate(cnt=Count('id'))
            .order_by('-cnt')[:1]
        )
        diag_str = diag_dist[0]['admission_diagnosis'][:50] if diag_dist else '—'

        row6 = [
            num, full_name,
            org['workplace_org__enterprise_code'] or '—',
            org['workplace_org__branch_code'] or '—',
            p_count, dept_str, diag_str, svc_total
        ]

        for col, val in enumerate(row6, 1):
            c = ws6.cell(row=r6, column=col, value=val)
            c.alignment = Alignment(
                horizontal='right' if col in (5, 8) else 'left',
                vertical='center', wrap_text=True
            )
            c.border = border
            if col == 8:
                c.number_format = '#,##0'
            if num % 2 == 0 and col < 6:
                c.fill = PatternFill('solid', fgColor='F8F9FA')

        ws6.row_dimensions[r6].height = 18
        r6 += 1

    # Jami
    ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=4)
    c = ws6.cell(row=r6, column=1, value="JAMI:")
    c.fill = header_fill; c.font = header_font; c.border = border

    c5 = ws6.cell(row=r6, column=5, value=grand_patients)
    c5.fill = header_fill; c5.font = header_font
    c5.alignment = Alignment(horizontal='right', vertical='center')
    c5.border = border

    ws6.merge_cells(start_row=r6, start_column=6, end_row=r6, end_column=7)
    c67 = ws6.cell(row=r6, column=6, value='')
    c67.fill = header_fill; c67.border = border

    c8 = ws6.cell(row=r6, column=8, value=grand_svc_total)
    c8.fill = header_fill; c8.font = header_font
    c8.number_format = '#,##0'
    c8.alignment = Alignment(horizontal='right', vertical='center')
    c8.border = border
    ws6.row_dimensions[r6].height = 25

    if r6 == 3:
        ws6.merge_cells('A3:H3')
        ws6.cell(row=3, column=1,
                 value="Tashkilotga biriktirilgan bemorlar yo'q").border = border

    # ==================== DORI SHEETI ====================
    from apps.services.models import PatientMedicine
    from django.db.models import Sum as MSum

    ws_med = wb.create_sheet("Dori-darmonlar")
    ws_med.column_dimensions['A'].width = 5
    ws_med.column_dimensions['B'].width = 40
    ws_med.column_dimensions['C'].width = 12
    ws_med.column_dimensions['D'].width = 12
    ws_med.column_dimensions['E'].width = 15
    ws_med.column_dimensions['F'].width = 20

    ws_med.merge_cells('A1:F1')
    c = ws_med.cell(row=1, column=1, value="DORI-DARMON STATISTIKASI")
    c.fill = header_fill; c.font = header_font
    c.alignment = center; c.border = border
    ws_med.row_dimensions[1].height = 28

    heads = ['№', 'Dori nomi', 'Birlik', 'Jami miqdor', 'Bemorlar soni', "Jami summa (so'm)"]
    for col, h in enumerate(heads, 1):
        c = ws_med.cell(row=2, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = center; c.border = border
    ws_med.row_dimensions[2].height = 24

    top_meds = (
        PatientMedicine.objects
        .filter(patient_card__in=qs)
        .values('medicine__name', 'medicine__unit')
        .annotate(total_qty=MSum('quantity'), total_sum=MSum('price'), cnt=Count('patient_card', distinct=True))
        .order_by('-total_sum')
    )

    med_grand = 0
    for ri, m in enumerate(top_meds, 1):
        row_data = [
            ri,
            m['medicine__name'],
            m['medicine__unit'],
            float(m['total_qty'] or 0),
            m['cnt'],
            float(m['total_sum'] or 0),
        ]
        med_grand += float(m['total_sum'] or 0)
        for col, val in enumerate(row_data, 1):
            c = ws_med.cell(row=ri+2, column=col, value=val)
            c.alignment = Alignment(horizontal='center' if col in (1,3,4,5) else ('right' if col==6 else 'left'), vertical='center')
            c.border = border
            if col == 6: c.number_format = '#,##0'
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='F8F9FA')
        ws_med.row_dimensions[ri+2].height = 18

    # Jami qator
    last_r = len(list(top_meds)) + 3
    ws_med.merge_cells(start_row=last_r, start_column=1, end_row=last_r, end_column=5)
    c = ws_med.cell(row=last_r, column=1, value="JAMI:")
    c.fill = header_fill; c.font = header_font; c.border = border
    c6 = ws_med.cell(row=last_r, column=6, value=med_grand)
    c6.fill = header_fill; c6.font = header_font
    c6.number_format = '#,##0'
    c6.alignment = Alignment(horizontal='right', vertical='center')
    c6.border = border
    ws_med.row_dimensions[last_r].height = 24

    # ==================== JAVOB ====================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bemorlar_royxati.xlsx"'
    wb.save(response)
    return response


def export_pdf(request):
    qs = get_filtered_queryset(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1 * cm, leftMargin=1 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'title', parent=styles['Heading1'],
        fontSize=14, alignment=1, spaceAfter=12
    )
    small_style = ParagraphStyle(
        'small', parent=styles['Normal'],
        fontSize=7, leading=9
    )

    elements = []
    elements.append(Paragraph(
        "Shifoxonadan chiqarilganlar statistik ro'yxati", title_style
    ))
    elements.append(Spacer(1, 0.3 * cm))

    table_headers = [
        "№", "Bayonnoma", "Ism-familiya", "Jins", "Tug'ilgan\nsana",
        "Bo'lim", "Yotqizilgan\nsana", "Yotgan\nkun",
        "Yakun", "Xulosa", "MKB-10", "Shifokor"
    ]

    table_data = [[Paragraph(h, small_style) for h in table_headers]]

    for i, patient in enumerate(qs, 1):
        row = [
            str(i),
            patient.medical_record_number or '',
            Paragraph(patient.full_name or '', small_style),
            patient.get_gender_display(),
            patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '',
            Paragraph(str(patient.department) if patient.department else '', small_style),
            patient.admission_date.strftime('%d.%m.%Y') if patient.admission_date else '',
            str(patient.days_in_hospital),
            patient.get_outcome_display(),
            str(patient.discharge_conclusion) if patient.discharge_conclusion else '—',
            patient.clinical_main_diagnosis or '',
            Paragraph(
                str(patient.attending_doctor) if patient.attending_doctor else '', small_style
            ),
        ]
        table_data.append(row)

    col_widths = [
        1 * cm, 2.5 * cm, 4.5 * cm, 1.2 * cm, 2.2 * cm,
        3.5 * cm, 2.5 * cm, 1.5 * cm, 2.5 * cm, 2.5 * cm, 1.8 * cm, 3.8 * cm
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
            colors.white, colors.HexColor('#F2F2F2')
        ]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bemorlar_royxati.pdf"'
    return response


# apps/statistic/exports.py — get_filtered_queryset yangilash
