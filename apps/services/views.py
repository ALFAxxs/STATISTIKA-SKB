# apps/services/views.py

import io
import json
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.utils import timezone

from apps.users.decorators import role_required
from apps.patients.models import PatientCard, Doctor
from .models import ServiceCategory, Service, PatientService
from .forms import PatientServiceForm, ServiceResultForm


# ==================== AJAX ====================

@login_required
def service_search(request):
    """AJAX — xizmat qidirish"""
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    patient_id = request.GET.get('patient_id', '')

    qs = Service.objects.filter(is_active=True).select_related('category').order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    if category_id:
        qs = qs.filter(category_id=category_id)
    # Bo'sh qidiruv — faqat 20 ta ko'rsatish
    # Kategoriya tanlangan bo'lsa — hammasi
    limit = 50 if category_id else 20

    # Bemor kategoriyasiga qarab narx hisoblash
    patient_category = 'railway'
    if patient_id:
        try:
            patient = PatientCard.objects.get(pk=patient_id)
            patient_category = patient.patient_category or 'railway'
        except PatientCard.DoesNotExist:
            pass

    data = []
    for s in qs[:limit]:
        price = s.price_for_patient(patient_category)
        data.append({
            'id': s.id,
            'name': str(s),
            'category': s.category.name,
            'category_id': s.category_id,
            'price': float(price),
            'price_normal': float(s.price_normal),
            'price_railway': float(s.price_railway),
        })

    return JsonResponse(data, safe=False)


# ==================== BEMOR XIZMATLARI ====================

@login_required
def patient_services(request, patient_pk):
    """Bemorning barcha xizmatlari"""
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    services = PatientService.objects.filter(
        patient_card=patient
    ).select_related(
        'service__category', 'ordered_by', 'performed_by'
    ).order_by('-ordered_at')

    categories = ServiceCategory.objects.filter(is_active=True)

    # Moliyaviy umumlama
    totals = services.aggregate(
        total_sum=Sum('price'),
        count=Count('id'),
    )
    total_price = (totals['total_sum'] or 0)

    # Kategoriya bo'yicha umumlama
    cat_stats = services.values(
        'service__category__name',
        'service__category__icon',
    ).annotate(
        count=Count('id'),
        total=Sum('price'),
    ).order_by('-total')

    return render(request, 'services/patient_services.html', {
        'patient': patient,
        'services': services,
        'categories': categories,
        'total_price': total_price,
        'total_count': totals['count'] or 0,
        'cat_stats': cat_stats,
    })


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception')
def add_service(request, patient_pk):
    """Bemorga xizmat qo'shish — AJAX"""
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service_id = data.get('service_id')
            quantity = int(data.get('quantity', 1))
            ordered_by_id = data.get('ordered_by_id')
            notes = data.get('notes', '')

            service = Service.objects.get(pk=service_id, is_active=True)
            price = service.price_for_patient(patient.patient_category or 'railway')

            ps = PatientService.objects.create(
                patient_card=patient,
                service=service,
                quantity=quantity,
                price=price,
                patient_category_at_order=patient.patient_category or 'railway',
                ordered_by_id=ordered_by_id if ordered_by_id else None,
                notes=notes,
            )

            return JsonResponse({
                'success': True,
                'id': ps.id,
                'service_name': service.name,
                'category': service.category.name,
                'quantity': quantity,
                'price': float(price),
                'total': float(ps.total_price),
                'status': ps.get_status_display(),
                'ordered_at': ps.ordered_at.strftime('%d.%m.%Y %H:%M'),
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'POST required'})


@login_required
@role_required('admin', 'doctor', 'statistician')
def update_service(request, pk):
    """Xizmat holatini yangilash — AJAX"""
    ps = get_object_or_404(PatientService, pk=pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            status = data.get('status')
            result = data.get('result', '')
            is_paid = data.get('is_paid', False)
            performed_by_id = data.get('performed_by_id')

            if status in dict(PatientService.STATUS_CHOICES):
                ps.status = status
            ps.result = result
            ps.is_paid = is_paid
            ps.performed_by_id = performed_by_id if performed_by_id else None

            if status == 'completed' and not ps.performed_at:
                ps.performed_at = timezone.now()

            ps.save()
            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False})


@login_required
@role_required('admin', 'doctor')
def delete_service(request, pk):
    """Xizmatni o'chirish"""
    ps = get_object_or_404(PatientService, pk=pk)
    patient_pk = ps.patient_card.pk

    if request.method == 'POST':
        if ps.status == 'ordered':
            ps.delete()
            return JsonResponse({'success': True})
        return JsonResponse({
            'success': False,
            'error': "Bajarilgan xizmatni o'chirib bo'lmaydi"
        })

    return JsonResponse({'success': False})


# ==================== STATISTIKA ====================

@login_required
@role_required('admin', 'statistician')
def service_statistics(request):
    """Xizmatlar statistikasi dashboard"""
    # Filterlar
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    category_id = request.GET.get('category', '')
    period = request.GET.get('period', 'month')  # day/month/year
    patient_cat = request.GET.get('patient_category', '')

    qs = PatientService.objects.exclude(status='cancelled').select_related(
        'service__category', 'patient_card'
    )

    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id:
        qs = qs.filter(service__category_id=category_id)
    if patient_cat:
        qs = qs.filter(patient_category_at_order=patient_cat)

    # Umumiy ko'rsatkichlar
    totals = qs.aggregate(
        total_sum=Sum('price'),
        count=Count('id'),
        railway_sum=Sum('price', filter=Q(patient_category_at_order='railway')),
        nonresident_sum=Sum('price', filter=Q(
            patient_category_at_order='non_resident'
        )),
    )

    # Kategoriya bo'yicha
    cat_stats = qs.values(
        'service__category__name',
        'service__category__icon',
        'service__category__id',
    ).annotate(
        count=Count('id'),
        total=Sum('price'),
    ).order_by('-total')

    # Eng ko'p ishlatiladigan xizmatlar (Top 10)
    top_services = qs.values(
        'service__name',
        'service__category__name',
    ).annotate(
        count=Count('id'),
        total=Sum('price'),
    ).order_by('-count')[:10]

    # Vaqt bo'yicha dinamika
    if period == 'day':
        time_stats = qs.annotate(
            period=TruncDay('ordered_at')
        ).values('period').annotate(
            count=Count('id'),
            total=Sum('price'),
        ).order_by('period')
    elif period == 'year':
        time_stats = qs.annotate(
            period=TruncYear('ordered_at')
        ).values('period').annotate(
            count=Count('id'),
            total=Sum('price'),
        ).order_by('period')
    else:
        time_stats = qs.annotate(
            period=TruncMonth('ordered_at')
        ).values('period').annotate(
            count=Count('id'),
            total=Sum('price'),
        ).order_by('period')

    time_labels = [
        item['period'].strftime('%Y-%m-%d' if period == 'day' else '%Y-%m' if period == 'month' else '%Y')
        for item in time_stats if item['period']
    ]
    time_counts = [item['count'] for item in time_stats if item['period']]
    time_totals = [float(item['total'] or 0) for item in time_stats if item['period']]

    categories = ServiceCategory.objects.filter(is_active=True)

    return render(request, 'services/statistics.html', {
        'totals': totals,
        'cat_stats': cat_stats,
        'top_services': top_services,
        'time_labels': json.dumps(time_labels),
        'time_counts': json.dumps(time_counts),
        'time_totals': json.dumps(time_totals),
        'cat_labels': json.dumps([c['service__category__name'] for c in cat_stats]),
        'cat_values': json.dumps([float(c['total'] or 0) for c in cat_stats]),
        'categories': categories,
        'date_from': date_from,
        'date_to': date_to,
        'selected_category': category_id,
        'selected_period': period,
        'selected_patient_cat': patient_cat,
        'current_filters': request.GET.urlencode(),
    })


# ==================== EXPORT ====================

@login_required
@role_required('admin', 'statistician')
def export_services_excel(request):
    """Xizmatlar hisobotini Excel ga export"""
    qs = PatientService.objects.exclude(status='cancelled').select_related(
        'service__category', 'patient_card', 'ordered_by', 'performed_by'
    )

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category_id = request.GET.get('category')
    patient_cat = request.GET.get('patient_category')

    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id:
        qs = qs.filter(service__category_id=category_id)
    if patient_cat:
        qs = qs.filter(patient_category_at_order=patient_cat)

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== 1-sahifa: Xizmatlar ro'yxati =====
    ws = wb.active
    ws.title = "Xizmatlar ro'yxati"

    headers = [
        '№', 'Sana', 'Bemor', 'Bemor kategoriyasi',
        'Kategoriya', 'Xizmat', 'Miqdori',
        'Narx', 'Jami', 'Holat', "To'langan",
        'Buyurtma bergan', 'Bajargan', 'Natija'
    ]
    ws.row_dimensions[1].height = 35
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    cat_display = {
        'railway': "Temir yo'lchi", 'paid': 'Pullik',
        'non_resident': 'Norezident', 'foreign': 'Chet el',
    }

    for i, ps in enumerate(qs.order_by('-ordered_at'), 1):
        row_data = [
            i,
            ps.ordered_at.strftime('%d.%m.%Y %H:%M'),
            ps.patient_card.full_name,
            cat_display.get(ps.patient_category_at_order, ps.patient_category_at_order),
            ps.service.category.name,
            ps.service.name,
            ps.quantity,
            float(ps.price),
            float(ps.total_price),
            ps.get_status_display(),
            'Ha' if ps.is_paid else "Yo'q",
            str(ps.ordered_by) if ps.ordered_by else '—',
            str(ps.performed_by) if ps.performed_by else '—',
            ps.result or '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
            if col == 11:
                cell.fill = PatternFill('solid', fgColor='C6EFCE' if ps.is_paid else 'FFC7CE')

    col_widths = [4, 16, 25, 16, 18, 30, 8, 12, 12, 14, 10, 22, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== 2-sahifa: Kategoriya statistikasi =====
    ws2 = wb.create_sheet("Kategoriyalar")
    ws2['A1'] = "Xizmat kategoriyalari bo'yicha statistika"
    ws2['A1'].font = Font(bold=True, size=13, color='1F4E79')

    h2 = ['Kategoriya', 'Xizmatlar soni', "Jami summa (so'm)", "To'langan (so'm)"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    cat_data = qs.values('service__category__name').annotate(
        count=Count('id'),
        total=Sum('price'),
    ).order_by('-total')

    for i, row in enumerate(cat_data, 4):
        ws2.cell(row=i, column=1, value=row['service__category__name']).border = border
        ws2.cell(row=i, column=2, value=row['count']).border = border
        ws2.cell(row=i, column=3, value=float(row['total'] or 0)).border = border

    for col, w in enumerate([25, 15, 20, 20], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ===== 3-sahifa: Umumiy hisobot =====
    ws3 = wb.create_sheet("Umumiy hisobot")
    ws3['A1'] = "Umumiy moliyaviy hisobot"
    ws3['A1'].font = Font(bold=True, size=13, color='1F4E79')

    totals_agg = qs.aggregate(
        total=Sum('price'),
        count=Count('id'),
        railway=Sum('price', filter=Q(patient_category_at_order='railway')),
        nonresident=Sum('price', filter=Q(
            patient_category_at_order='non_resident'
        )),
    )
    summary = [
        ('Jami xizmatlar soni', totals_agg['count'] or 0),
        ("Jami summa (so'm)", float(totals_agg['total'] or 0)),
        ("Temir yo'lchilar daromadi (so'm)", float(totals_agg['railway'] or 0)),
        ("Norezidentlar daromadi (so'm)", float(totals_agg['nonresident'] or 0)),
    ]
    for i, (label, val) in enumerate(summary, 3):
        ws3.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=val)

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="xizmatlar_hisoboti.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required('admin', 'statistician')
def export_services_pdf(request):
    """Xizmatlar hisobotini PDF ga export"""
    qs = PatientService.objects.exclude(status='cancelled').select_related(
        'service__category', 'patient_card'
    ).order_by('-ordered_at')

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1*cm
    )
    styles = getSampleStyleSheet()
    small = ParagraphStyle('sm', parent=styles['Normal'], fontSize=7, leading=9)
    title_style = ParagraphStyle(
        'title', parent=styles['Heading1'],
        fontSize=14, alignment=1, spaceAfter=10
    )

    elements = [
        Paragraph("Xizmatlar hisoboti", title_style),
        Spacer(1, 0.3*cm)
    ]

    headers = [
        '№', 'Sana', 'Bemor', 'Kategoriya',
        'Xizmat', 'Miqdor', 'Narx', 'Jami', "To'langan"
    ]
    table_data = [[Paragraph(h, small) for h in headers]]

    for i, ps in enumerate(qs, 1):
        table_data.append([
            str(i),
            ps.ordered_at.strftime('%d.%m.%Y'),
            Paragraph(ps.patient_card.full_name, small),
            Paragraph(ps.service.category.name, small),
            Paragraph(ps.service.name, small),
            str(ps.quantity),
            f"{float(ps.price):,.0f}",
            f"{float(ps.total_price):,.0f}",
            'Ha' if ps.is_paid else "Yo'q",
        ])

    col_widths = [
        1*cm, 2.5*cm, 4*cm, 3*cm,
        5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2*cm
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="xizmatlar_hisoboti.pdf"'
    return response